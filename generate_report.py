import os
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials

# ── 1. Compute Mon-Fri dates for the current week ──────────────────────────
today = datetime.today()
monday = today - timedelta(days=today.weekday())
week_dates = [monday + timedelta(days=i) for i in range(5)]

week_label  = f"{monday.strftime('%B %d')} - {week_dates[4].strftime('%d, %Y')}"
date_header = f"{monday.strftime('%B %d')} - {week_dates[4].strftime('%d, %Y')} "

print(f"=== Generating report for week: {week_label} ===")
print(f"Target dates: {[d.strftime('%A %b %d') for d in week_dates]}")
print(f"Target day numbers: {[d.day for d in week_dates]}")

github_env = os.environ.get("GITHUB_ENV", "")
if github_env:
    with open(github_env, "a") as f:
        f.write(f"WEEK_LABEL={week_label}\n")

# ── 2. Connect to Google Sheets and read leave data ────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

creds_json     = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
spreadsheet_id = os.environ.get("SPREADSHEET_ID", "")

leave_map = {}

def parse_leave_type(text, name_part):
    rest = text[len(name_part):].strip(" \t-–()")
    rest_up = rest.upper()
    if not rest_up or rest_up == "ON LEAVE":
        return "On Leave"
    elif "EMERGENCY LEAVE" in rest_up or re.search(r'\bEL\b', rest_up):
        return "Emergency Leave"
    elif re.search(r'\bSL\b', rest_up) or "SICK" in rest_up:
        return "Sick Leave"
    elif re.search(r'\bHL\b', rest_up) or "HALF" in rest_up:
        return "Half Day"
    else:
        return "On Leave"

print(f"\n=== Google Sheets Connection ===")
print(f"GOOGLE_CREDENTIALS_JSON present: {bool(creds_json)}")
print(f"SPREADSHEET_ID present: {bool(spreadsheet_id)}")
print(f"SPREADSHEET_ID value: {spreadsheet_id}")

if creds_json and spreadsheet_id:
    try:
        creds_info = json.loads(creds_json)
        print(f"Service account email: {creds_info.get('client_email', 'NOT FOUND')}")

        creds = Credentials.from_service_account_info(creds_info, scopes=SCOPES)
        gc    = gspread.authorize(creds)
        print("✅ Google auth successful")

        month_tab = monday.strftime("%B %Y").upper()
        print(f"Looking for tab: '{month_tab}'")

        wb_gs = gc.open_by_key(spreadsheet_id)
        print(f"✅ Spreadsheet opened: {wb_gs.title}")

        # List all available tabs
        all_tabs = [ws.title for ws in wb_gs.worksheets()]
        print(f"Available tabs: {all_tabs}")

        try:
            ws_gs = wb_gs.worksheet(month_tab)
            print(f"✅ Tab '{month_tab}' found!")
        except gspread.exceptions.WorksheetNotFound:
            ws_gs = wb_gs.get_worksheet(0)
            print(f"⚠️ Tab '{month_tab}' NOT found, using first sheet: '{ws_gs.title}'")

        all_values = ws_gs.get_all_values()
        print(f"Total rows read: {len(all_values)}")

        # Print first 5 rows for debug
        print("First 3 rows of sheet:")
        for i, row in enumerate(all_values[:3]):
            print(f"  Row {i+1}: {row}")

        # Map column index → weekday index
        header_row = all_values[1] if len(all_values) > 1 else []
        print(f"\nHeader row (row 2): {header_row}")

        col_to_weekday = {}
        for ci, hdr in enumerate(header_row):
            h = hdr.strip().upper()
            if   h == "MONDAY":    col_to_weekday[ci] = 0
            elif h == "TUESDAY":   col_to_weekday[ci] = 1
            elif h == "WEDNESDAY": col_to_weekday[ci] = 2
            elif h == "THURSDAY":  col_to_weekday[ci] = 3
            elif h == "FRIDAY":    col_to_weekday[ci] = 4

        print(f"Column to weekday map: {col_to_weekday}")

        target_days = {wd.day: idx for idx, wd in enumerate(week_dates)}
        print(f"Target day numbers → index: {target_days}")

        print("\nScanning calendar rows...")
        for row_idx, row in enumerate(all_values[2:], start=3):
            for ci, cell_text in enumerate(row):
                cell_text = cell_text.strip()
                if not cell_text or ci not in col_to_weekday:
                    continue

                lines = [l.strip() for l in cell_text.splitlines() if l.strip()]
                if not lines:
                    continue

                # Try to parse date number from first line
                date_num  = None
                name_lines = lines
                try:
                    date_num   = int(lines[0])
                    name_lines = lines[1:]
                except ValueError:
                    pass

                # Debug: print cells that have date numbers matching our week
                if date_num in target_days:
                    print(f"  ✅ Row {row_idx}, Col {ci}: date={date_num}, names={name_lines}")

                if date_num not in target_days:
                    continue

                day_idx = target_days[date_num]

                for nl in name_lines:
                    if not nl:
                        continue
                    name_match = re.match(r"^([A-Za-z]+)", nl)
                    if not name_match:
                        continue
                    raw_name   = name_match.group(1).strip()
                    leave_type = parse_leave_type(nl, raw_name)

                    if raw_name not in leave_map:
                        leave_map[raw_name] = {}
                    leave_map[raw_name][day_idx] = leave_type
                    print(f"  📌 {raw_name} | {week_dates[day_idx].strftime('%A %b %d')} | {leave_type}")

    except Exception as e:
        import traceback
        print(f"❌ ERROR: {e}")
        print(traceback.format_exc())
else:
    if not creds_json:
        print("❌ GOOGLE_CREDENTIALS_JSON is empty or missing!")
    if not spreadsheet_id:
        print("❌ SPREADSHEET_ID is empty or missing!")

print(f"\n=== Final leave_map: {leave_map} ===\n")

# ── 3. Build the Excel report ──────────────────────────────────────────────
wb = load_workbook("template_base.xlsx")

sheets_config = [
    ("SEO Template", ["James", "Jone", "Bernard", "Ken"]),
    ("RTEMPLATE",    ["Rosee", "Nicole"]),
    ("ADTemplate",   ["Adolf"]),
    ("DSTEMPLATE",   ["Kae", "Charles"]),
]

DAY_COL = {0: "B", 1: "C", 2: "D", 3: "E", 4: "F"}

LEAVE_COLORS = {
    "On Leave":        "D9E1F2",
    "Sick Leave":      "FFD9D9",
    "Half Day":        "FFF2CC",
    "Emergency Leave": "F4CCFF",
}

for template_name, names in sheets_config:
    for name in names:
        new_ws = wb.copy_worksheet(wb[template_name])
        new_ws.title = name
        new_ws["A10"] = name
        new_ws["C3"]  = date_header

        if template_name == "SEO Template":
            date_cells = ["L13", "L14", "L15", "L16", "L17"]
        else:
            date_cells = ["L8", "L9", "L10", "L11", "L12"]

        for i, ca in enumerate(date_cells):
            new_ws[ca] = week_dates[i]

        new_ws.column_dimensions["A"].width = 28

        person_leaves = {}
        for lname, ldays in leave_map.items():
            if lname.lower() == name.lower():
                person_leaves = ldays
                break

        for day_idx, leave_type in person_leaves.items():
            col = DAY_COL.get(day_idx)
            if not col:
                continue
            cell = new_ws[f"{col}13"]
            cell.value = leave_type
            color = LEAVE_COLORS.get(leave_type, "D9E1F2")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True)
            print(f"  ✅ Marked {name} → {col}13 = '{leave_type}'")

wb.save("weekly_report.xlsx")
print("\n✅ Report saved as weekly_report.xlsx")
